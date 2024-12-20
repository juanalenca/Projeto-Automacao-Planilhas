import tkinter as tk
from tkinter import filedialog, messagebox
import re
import os
import pandas as pd
import sqlite3
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# Definir variáveis globais
caminho_pasta = 'C:\\AutomacaoTabelasReciprev\\'  # Caminho base do projeto
diretorio_planilhas = os.path.join(caminho_pasta, 'arquivos-analise\\planilhas_geradas')  # Diretório para salvar as planilhas geradas
caminho_banco_dados = os.path.join(caminho_pasta, 'contab_reciprev.db')  # Caminho para o banco de dados SQLite
nome_aba_planilha_dinamica = '2020 DADOS'  # Nome da aba da planilha dinâmica existente
nome_aba_calculo_final = 'CÁLCULO FINAL FOLHA'  # Nome da aba onde será inserido o cálculo final
dynamic_spreadsheet_path = ''  # Caminho para a planilha dinâmica selecionada pelo usuário

def processar_arquivo(caminho_arquivo, diretorio_planilhas, dynamic_spreadsheet_path):
    """
    Função principal para processar um arquivo de texto e atualizar a planilha dinâmica e o banco de dados.
    """
    # Extrair o nome do arquivo sem extensão
    nome_arquivo = os.path.splitext(os.path.basename(caminho_arquivo))[0]

    # Conectar ao banco de dados SQLite
    conexao = sqlite3.connect(caminho_banco_dados)
    cursor = conexao.cursor()

    # Carregar a planilha dinâmica existente
    workbook = load_workbook(dynamic_spreadsheet_path)
    worksheet = workbook[nome_aba_planilha_dinamica]

    # Verificar se a aba "CÁLCULO FINAL FOLHA" já existe; se não, criar uma nova
    if nome_aba_calculo_final in workbook.sheetnames:
        worksheet_calculo_final = workbook[nome_aba_calculo_final]
    else:
        worksheet_calculo_final = workbook.create_sheet(nome_aba_calculo_final)

    # Ler o conteúdo do arquivo de texto
    with open(caminho_arquivo, 'r', encoding='utf-8') as file:
        conteudo = file.read()

    # Definir padrões regex para extrair as seções de vantagens e descontos
    padrao_vantagens = r'(COD\s+V A N T A G E N S\s+TOT\. FUNC\.\s+TOT\. VALOR\s+COD\s+V A N T A G E N S\s+TOT\. FUNC\.\s+TOT\. VALOR[\s\S]+?)(?=\nCOD\s+D E S C O N T O S\s+TOT\. FUNC\.\s+TOT\. VALOR|\Z)'
    padrao_descontos = r'(COD\s+D E S C O N T O S\s+TOT\. FUNC\.\s+TOT\. VALOR\s+COD\s+D E S C O N T O S\s+TOT\. FUNC\.\s+TOT\. VALOR[\s\S]+?)(?=\nTOTAL\s+LIQUIDO|\Z)'

    def salvar_parte(texto, parte):
        """
        Salva uma parte do texto em um arquivo .txt para análise ou registro.
        """
        caminho = os.path.join(diretorio_planilhas, f'{nome_arquivo}_parte{parte}.txt')
        with open(caminho, 'w', encoding='utf-8') as f:
            f.write(texto.strip())

    def dividir_tabela(texto):
        """
        Divide a tabela em duas partes, pois os dados estão dispostos em duas colunas no arquivo de texto.
        """
        linhas = texto.strip().split('\n')
        metade_index = len(linhas[0]) // 2  # Índice do meio da linha
        cabecalho1, cabecalho2 = linhas[0][:metade_index].strip(), linhas[0][metade_index:].strip()
        linha_separadora = '=' * len(cabecalho1)  # Linha separadora para formatação

        dados1, dados2 = [cabecalho1, linha_separadora], [cabecalho2, linha_separadora]
        for linha in linhas[2:]:
            if len(linha) > metade_index:
                parte1, parte2 = linha[:metade_index].strip(), linha[metade_index:].strip()
            else:
                parte1, parte2 = linha.strip(), ""
            dados1.append(parte1)
            dados2.append(parte2)

        tabela1 = '\n'.join(dados1)
        tabela2 = '\n'.join(dados2)
        return tabela1, tabela2

    def limpar_cod(cod):
        """
        Remove caracteres não numéricos do código.
        """
        return re.sub(r'\D', '', cod).strip()

    def validar_numero(valor):
        """
        Converte uma string numérica formatada em um número float.
        """
        try:
            valor = valor.replace('.', '').replace(',', '.')
            return float(valor)
        except ValueError:
            return 0.0

    def obter_classificacao(cursor, verba):
        """
        Obtém a classificação de uma verba a partir do banco de dados.
        """
        # Primeiro, obter o id da verba na tabela 'verbas'
        cursor.execute("SELECT id FROM verbas WHERE verba = ?", (verba,))
        verba_id = cursor.fetchone()

        if verba_id:
            # Usar o verba_id para buscar a classificação na tabela 'classificacoes'
            cursor.execute("SELECT classificacao FROM classificacoes WHERE verba_id = ?", (verba_id[0],))
            resultado = cursor.fetchone()
            return resultado[0] if resultado else None
        else:
            # Se não encontrar a verba, retornar None
            return None

    def extrair_dados(texto, tipo, categoria, orgao):
        """
        Extrai os dados das tabelas de vantagens ou descontos e retorna uma lista de tuplas com os dados.
        """
        linhas = texto.strip().split('\n')[2:]  # Ignorar cabeçalhos
        dados = []

        for linha in linhas:
            # Ignorar linhas vazias ou linhas que não parecem conter dados válidos
            if not linha.strip() or len(linha.strip()) < 20:
                continue

            campos = re.split(r'\s{2,}', linha.strip())

            # Verifica se há ao menos 4 campos antes de tentar processar a linha
            if len(campos) >= 4:
                cod = limpar_cod(campos[0])
                descricao = campos[1].strip()
                tot_func = validar_numero(campos[2])
                tot_valor = validar_numero(campos[3])
                classificacao = obter_classificacao(cursor, cod)

                # Criar a tupla de dados com os 8 elementos necessários
                dado = (tipo, cod, descricao, tot_func, tot_valor, classificacao, categoria, orgao)

                # Se a tupla contém exatamente 8 elementos, ela é considerada válida
                if len(dado) == 8:
                    dados.append(dado)
                else:
                    print(f"Erro: Tupla com estrutura incorreta: {dado}")
            else:
                print(f"Erro: Linha com campos insuficientes: {linha}")

        return dados

    def salvar_planilha(dados, nome_planilha):
        """
        Salva os dados extraídos em uma planilha Excel para registro.
        """
        df = pd.DataFrame(dados, columns=['TIPO', 'COD', 'DESCRIÇÃO', 'TOT. FUNC.', 'TOT. VALOR', 'CLASSIFICAÇÃO', 'CATEGORIA', 'ORGÃO'])
        caminho_planilha = os.path.join(diretorio_planilhas, f'{nome_planilha}.xlsx')
        df.to_excel(caminho_planilha, index=False)

    def inserir_dados_planilha_dinamica(dados, worksheet):
        """
        Insere os dados na planilha dinâmica existente, destacando as novas linhas em amarelo.
        """
        fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # Amarelo

        # Encontrar a última linha com dados
        ultima_linha = worksheet.max_row
        while not any(cell.value for cell in worksheet[ultima_linha]):
            ultima_linha -= 1

        # Inserir dados logo abaixo da última linha com dados
        for dado in dados:
            worksheet.append(dado)
            ultima_linha += 1
            for cell in worksheet[ultima_linha]:
                cell.fill = fill  # Destacar a nova linha

    def atualizar_banco_de_dados(dados):
        """
        Atualiza o banco de dados SQLite com os novos dados.
        """
        try:
            conexao = sqlite3.connect(caminho_banco_dados)
            cursor = conexao.cursor()

            for dado in dados:
                tipo, cod, descricao, tot_func, tot_valor, classificacao, categoria, orgao = dado

                # Verificar se a verba já existe
                cursor.execute('''
                    SELECT id FROM verbas WHERE verba = ?
                ''', (cod,))
                resultado = cursor.fetchone()

                if resultado:
                    verba_id = resultado[0]
                    print(f"Verba já existe: {cod} - Ignorando inserção.")
                else:
                    # Inserir na tabela 'verbas'
                    cursor.execute('''
                        INSERT INTO verbas (verba, tipo, descricao, quantidade, total_valor)
                        VALUES (?, ?, ?, ?, ?)
                    ''', (cod, tipo, descricao, tot_func, tot_valor))
                    verba_id = cursor.lastrowid

                # Inserir na tabela 'classificacoes'
                if classificacao:
                    cursor.execute('''
                        INSERT INTO classificacoes (verba_id, classificacao)
                        VALUES (?, ?)
                    ''', (verba_id, classificacao))

                # Inserir na tabela 'orgaos'
                if orgao:
                    cursor.execute('''
                        INSERT INTO orgaos (verba_id, orgao)
                        VALUES (?, ?)
                    ''', (verba_id, orgao))

                # Inserir na tabela 'categorias'
                if categoria:
                    cursor.execute('''
                        INSERT INTO categorias (verba_id, categoria)
                        VALUES (?, ?)
                    ''', (verba_id, categoria))

                # Inserir na tabela 'codigo_orgaos'
                if cod:
                    cursor.execute('''
                        INSERT INTO codigo_orgaos (verba_id, codigo_orgao)
                        VALUES (?, ?)
                    ''', (verba_id, cod))

            conexao.commit()
        except sqlite3.Error as e:
            print(f"Erro ao inserir dados no banco de dados: {e}")
        finally:
            conexao.close()

    def extrair_categoria(conteudo):
        """
        Extrai a categoria (APO, PEN ou PPR) do conteúdo do arquivo de texto.
        """
        padrao_categoria = r'\b(APO|PEN|PPR)\b'
        match = re.search(padrao_categoria, conteudo)
        if match:
            return match.group(0)
        else:
            print("Aviso: Categoria não encontrada no conteúdo.")
            return "CATEGORIA NÃO ENCONTRADA"

    def extrair_orgao(conteudo):
        """
        Extrai o órgão específico do conteúdo do arquivo de texto.
        """
        padroes_orgao = [
            r'\bINATIVOS E PENSIONISTAS CAMARA\b',
            r'\bINATIVOS E PENSIONISTAS FUNDACAO CULTURA\b',
            r'\bINATIVOS E PENSIONISTAS GERALDAO\b',
            r'\bINATIVOS E PENSIONISTAS SETOR EDUCACIONA\b',
            r'\bINATIVOS E PENSIONS. SIST PREVIDENCIARIO\b',
            r'\bINATIVOS E PENSIONISTAS IASC\b'
        ]
        for padrao in padroes_orgao:
            match = re.search(padrao, conteudo)
            if match:
                return match.group(0)
        print("Aviso: Órgão não encontrado no conteúdo.")
        return "ÓRGÃO NÃO ENCONTRADO"

    def inserir_dados_calculo_final(dados, worksheet_calculo_final):
        """
        Insere os dados na aba 'CÁLCULO FINAL FOLHA' da planilha, estruturando-os de forma hierárquica.
        """
        # Agrupar dados por 'TIPO', 'CATEGORIA' e 'ORGÃO'
        dados_agrupados = {}
        for dado in dados:
            if len(dado) >= 8:
                tipo_categoria_orgao = (dado[0], dado[6], dado[7])  # ('TIPO', 'CATEGORIA', 'ORGÃO')
                if tipo_categoria_orgao not in dados_agrupados:
                    dados_agrupados[tipo_categoria_orgao] = []
                dados_agrupados[tipo_categoria_orgao].append(dado)
            else:
                print(f"Erro: Tupla com estrutura incorreta: {dado}")

        # Definir estilos para formatação
        header_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')  # Azul claro
        bold_font = Font(bold=True)
        alignment_center = Alignment(horizontal='center')
        alignment_left = Alignment(horizontal='left')
        alignment_right = Alignment(horizontal='right')

        # Inserir os dados na planilha de forma hierárquica
        linha_atual = 1  # Controla a linha atual para inserção de dados
        for (tipo, categoria, orgao), items in dados_agrupados.items():
            # Inserir título do tipo e órgão
            worksheet_calculo_final[f'A{linha_atual}'] = f'{tipo} - {orgao}'
            worksheet_calculo_final[f'A{linha_atual}'].font = bold_font
            worksheet_calculo_final[f'A{linha_atual}'].alignment = alignment_left
            worksheet_calculo_final.merge_cells(f'A{linha_atual}:E{linha_atual}')
            linha_atual += 1

            # Agrupar por classificação
            classificacao_grupos = {}
            for item in items:
                classificacao = item[5]
                if classificacao not in classificacao_grupos:
                    classificacao_grupos[classificacao] = []
                classificacao_grupos[classificacao].append(item)

            # Inserir dados para cada classificação
            for classificacao, itens_classificados in classificacao_grupos.items():
                total_valor_classificacao = sum([item[4] for item in itens_classificados])
                worksheet_calculo_final[f'B{linha_atual}'] = classificacao
                worksheet_calculo_final[f'C{linha_atual}'] = f'R$ {total_valor_classificacao:,.2f}'
                worksheet_calculo_final[f'B{linha_atual}'].font = bold_font
                worksheet_calculo_final[f'C{linha_atual}'].font = bold_font
                worksheet_calculo_final[f'B{linha_atual}'].alignment = alignment_left
                worksheet_calculo_final[f'C{linha_atual}'].alignment = alignment_left

                linha_atual += 1

                for item in itens_classificados:
                    cod = item[1]
                    descricao = item[2]
                    tot_valor = item[4]
                    worksheet_calculo_final[f'C{linha_atual}'] = cod
                    worksheet_calculo_final[f'D{linha_atual}'] = descricao
                    worksheet_calculo_final[f'E{linha_atual}'] = f'R$ {tot_valor:,.2f}'
                    worksheet_calculo_final[f'C{linha_atual}'].alignment = alignment_center
                    worksheet_calculo_final[f'D{linha_atual}'].alignment = alignment_left
                    worksheet_calculo_final[f'E{linha_atual}'].alignment = alignment_right

                    linha_atual += 1

                # Adicionar uma linha em branco para separar grupos
                linha_atual += 1

        # Ajustar a largura das colunas
        for col in range(1, 6):
            worksheet_calculo_final.column_dimensions[get_column_letter(col)].width = 20

    # Extrair a categoria e o órgão do conteúdo do arquivo
    categoria = extrair_categoria(conteudo)
    orgao = extrair_orgao(conteudo)

    # Processar as seções de vantagens
    vantagens = re.findall(padrao_vantagens, conteudo)
    for i, vantagem in enumerate(vantagens):
        # Dividir a tabela em duas partes (colunas)
        parte1, parte2 = dividir_tabela(vantagem)
        salvar_parte(parte1, f'vantagens_{i + 1}_parte1')
        salvar_parte(parte2, f'vantagens_{i + 1}_parte2')

        # Extrair dados de cada parte
        dados_parte1 = extrair_dados(parte1, 'P', categoria, orgao)
        dados_parte2 = extrair_dados(parte2, 'P', categoria, orgao)

        # Salvar dados em planilhas individuais
        salvar_planilha(dados_parte1, f'vantagens_{i + 1}_parte1')
        salvar_planilha(dados_parte2, f'vantagens_{i + 1}_parte2')

        # Inserir dados na planilha dinâmica
        inserir_dados_planilha_dinamica(dados_parte1, worksheet)
        inserir_dados_planilha_dinamica(dados_parte2, worksheet)

        # Atualizar o banco de dados
        atualizar_banco_de_dados(dados_parte1)
        atualizar_banco_de_dados(dados_parte2)

        # Inserir dados na aba de cálculo final
        inserir_dados_calculo_final(dados_parte1, worksheet_calculo_final)
        inserir_dados_calculo_final(dados_parte2, worksheet_calculo_final)

    # Processar as seções de descontos
    descontos = re.findall(padrao_descontos, conteudo)
    for i, desconto in enumerate(descontos):
        # Dividir a tabela em duas partes (colunas)
        parte1, parte2 = dividir_tabela(desconto)
        salvar_parte(parte1, f'descontos_{i + 1}_parte1')
        salvar_parte(parte2, f'descontos_{i + 1}_parte2')

        # Extrair dados de cada parte
        dados_parte1 = extrair_dados(parte1, 'D', categoria, orgao)
        dados_parte2 = extrair_dados(parte2, 'D', categoria, orgao)

        # Salvar dados em planilhas individuais
        salvar_planilha(dados_parte1, f'descontos_{i + 1}_parte1')
        salvar_planilha(dados_parte2, f'descontos_{i + 1}_parte2')

        # Inserir dados na planilha dinâmica
        inserir_dados_planilha_dinamica(dados_parte1, worksheet)
        inserir_dados_planilha_dinamica(dados_parte2, worksheet)

        # Atualizar o banco de dados
        atualizar_banco_de_dados(dados_parte1)
        atualizar_banco_de_dados(dados_parte2)

        # Inserir dados na aba de cálculo final
        inserir_dados_calculo_final(dados_parte1, worksheet_calculo_final)
        inserir_dados_calculo_final(dados_parte2, worksheet_calculo_final)

    # Fechar a conexão com o banco de dados
    conexao.close()

    # Salvar as alterações na planilha dinâmica
    workbook.save(dynamic_spreadsheet_path)

def anexar_arquivos():
    """
    Função para anexar arquivos de texto (.txt) ao aplicativo.
    """
    arquivos = filedialog.askopenfilenames(filetypes=[("Text Files", "*.txt")])
    for arquivo in arquivos:
        lista_arquivos.insert(tk.END, arquivo)

def anexar_planilha_dinamica():
    """
    Função para anexar a planilha dinâmica existente que será atualizada.
    """
    global dynamic_spreadsheet_path
    dynamic_spreadsheet_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx")])
    if dynamic_spreadsheet_path:
        label_planilha_dinamica.config(text=os.path.basename(dynamic_spreadsheet_path))

def gerar_planilha_dinamica():
    """
    Função que inicia o processamento dos arquivos anexados e atualiza a planilha dinâmica.
    """
    if not dynamic_spreadsheet_path:
        messagebox.showwarning("Aviso", "Por favor, anexe uma planilha dinâmica.")
        return

    if lista_arquivos.size() == 0:
        messagebox.showwarning("Aviso", "Por favor, anexe arquivos de texto para processar.")
        return

    for idx in range(lista_arquivos.size()):
        caminho_arquivo = lista_arquivos.get(idx)
        processar_arquivo(caminho_arquivo, diretorio_planilhas, dynamic_spreadsheet_path)

    messagebox.showinfo("Sucesso", "A planilha dinâmica foi atualizada com sucesso.")
    lista_arquivos.delete(0, tk.END)
    label_planilha_dinamica.config(text="Nenhum arquivo anexado")

# Configuração da interface gráfica (GUI) usando tkinter
root = tk.Tk()
root.title("Processador de Arquivos")

# Frame para anexar arquivos de texto
frame_arquivos = tk.Frame(root)
frame_arquivos.pack(pady=10)

label_arquivos = tk.Label(frame_arquivos, text="Arquivos de texto:")
label_arquivos.pack(side=tk.LEFT)

btn_anexar_arquivos = tk.Button(frame_arquivos, text="Anexar Arquivos", command=anexar_arquivos)
btn_anexar_arquivos.pack(side=tk.LEFT, padx=5)

# Lista de arquivos anexados
lista_arquivos = tk.Listbox(root, width=80, height=10)
lista_arquivos.pack(pady=10)

# Frame para anexar a planilha dinâmica
frame_planilha_dinamica = tk.Frame(root)
frame_planilha_dinamica.pack(pady=10)

label_planilha_dinamica = tk.Label(frame_planilha_dinamica, text="Nenhum arquivo anexado")
label_planilha_dinamica.pack(side=tk.LEFT)

btn_anexar_planilha_dinamica = tk.Button(frame_planilha_dinamica, text="Anexar Planilha Dinâmica", command=anexar_planilha_dinamica)
btn_anexar_planilha_dinamica.pack(side=tk.LEFT, padx=5)

# Botão para gerar a planilha dinâmica
btn_gerar_planilha_dinamica = tk.Button(root, text="Gerar Planilha Dinâmica", command=gerar_planilha_dinamica)
btn_gerar_planilha_dinamica.pack(pady=20)

# Iniciar o loop principal da interface gráfica
root.mainloop()
